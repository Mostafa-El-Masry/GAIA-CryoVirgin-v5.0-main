import sys
import csv
import re
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell


def read_csv_rows(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def extract_first_date_from_details(detail_rows):
    for row in detail_rows:
        value = row.get("Answered", "") or ""
        m = re.search(r"(\d{2}/\d{2}/\d{4})", value)
        if m:
            return m.group(1)
    return None


def normalize_detail_rows(detail_rows):
    """
    If Abandoned is NONE but status shows Abandoned(Handled/Call Center(887/888)),
    change Abandoned to Call Center<887> / Call Center<888> so they are credited
    properly in counts.
    """
    out = []
    for row in detail_rows:
        new_row = dict(row)
        abandoned = (new_row.get("Abandoned") or "").strip().upper()
        status = (new_row.get("AVG Handle Time") or "").strip()

        handled_887 = re.search(r"Handled\s*/\s*Call Center\(887\)", status, re.IGNORECASE)
        handled_888 = re.search(r"Handled\s*/\s*Call Center\(888\)", status, re.IGNORECASE)

        if abandoned == "NONE" and handled_887:
            new_row["Abandoned"] = "Call Center<887>"
        elif abandoned == "NONE" and handled_888:
            new_row["Abandoned"] = "Call Center<888>"

        out.append(new_row)
    return out


def compute_call_center_counts(detail_rows):
    """
    Count calls that should be credited to Call Center 887 / 888.
    """
    sarah = 0
    steffi = 0

    for row in detail_rows:
        agent = (row.get("Abandoned") or "").strip()
        status = (row.get("AVG Handle Time") or "").strip()

        is_cc_887 = agent == "Call Center<887>" or re.search(r"Call Center\(887\)", status)
        is_cc_888 = agent == "Call Center<888>" or re.search(r"Call Center\(888\)", status)

        is_answered = (
            status == "Answered"
            or re.search(r"Abandoned\(Handled\s*/\s*Call Center\(887\)", status)
            or re.search(r"Abandoned\(Handled\s*/\s*Call Center\(888\)", status)
        )

        if not is_answered:
            continue

        if is_cc_887:
            sarah += 1
        elif is_cc_888:
            steffi += 1

    return sarah, steffi


def set_cell_value(ws, row, column, value):
    """
    Safely set a cell value, skipping cells that are part of a merged range
    but are not the top-left anchor (those are read-only in openpyxl).
    """
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def fill_summary_template(summary_csv, detail_csv, template_path, output_path):
    summary_rows = read_csv_rows(summary_csv)
    raw_detail_rows = read_csv_rows(detail_csv)
    detail_rows = normalize_detail_rows(raw_detail_rows)

    if not summary_rows or not detail_rows:
        raise RuntimeError("One or both CSV files are empty or invalid.")

    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    # Title
    date_label = extract_first_date_from_details(detail_rows)
    if date_label:
        title = f"Call Center-Report-{date_label}"
    else:
        label = Path(summary_csv).stem
        title = f"Call Center-Report-{label}"

    set_cell_value(ws, 1, 1, title)

    # Separate branches and total
    branches = []
    total_row = None
    for row in summary_rows:
        q = (row.get("Queue") or "").strip()
        if not q:
            continue
        lower = q.lower()
        if lower == "total":
            total_row = row
        else:
            branches.append(row)

    sarah_count, steffi_count = compute_call_center_counts(detail_rows)

    row_index = 3  # Excel row 3
    sl = 1
    total_calls_from_branches = 0

    for row in branches:
        total_calls = int(row.get("Total Calls") or 0)
        answered = int(row.get("Answered") or 0)
        missed = int(row.get("Missed") or 0)
        abandoned = int(row.get("Abandoned") or 0)
        missed_abandoned = missed + abandoned

        answered_rate = row.get("Answered Rate") or ""
        abandon_rate = row.get("Abandon Rate") or ""

        set_cell_value(ws, row_index, 1, sl)
        set_cell_value(ws, row_index, 2, row.get("Queue") or "")
        set_cell_value(ws, row_index, 3, total_calls)
        set_cell_value(ws, row_index, 4, answered)
        set_cell_value(ws, row_index, 5, missed_abandoned)
        set_cell_value(ws, row_index, 6, row.get("AVG Handle Time") or "")
        set_cell_value(ws, row_index, 7, row.get("AVG Waiting Time (Answered Calls)") or "")
        set_cell_value(ws, row_index, 8, row.get("AVG Waiting Time (All Calls)") or "")
        set_cell_value(ws, row_index, 9, row.get("Max Waiting Time (All Calls)") or "")
        set_cell_value(ws, row_index, 10, row.get("Average Talking Time") or "")
        set_cell_value(ws, row_index, 11, answered_rate or None)
        set_cell_value(ws, row_index, 12, abandon_rate or None)
        set_cell_value(ws, row_index, 13, None)  # Sales Rate left for manual input

        total_calls_from_branches += total_calls
        row_index += 1
        sl += 1

    # Call Center <887-Sara>
    set_cell_value(ws, row_index, 1, sl)
    set_cell_value(ws, row_index, 2, "Call Center <887-Sara>")
    set_cell_value(ws, row_index, 3, sarah_count)
    for col in range(4, 14):
        set_cell_value(ws, row_index, col, None)
    row_index += 1
    sl += 1

    # Call Center <888-Sansa>
    set_cell_value(ws, row_index, 1, sl)
    set_cell_value(ws, row_index, 2, "Call Center <888-Sansa>")
    set_cell_value(ws, row_index, 3, steffi_count)
    for col in range(4, 14):
        set_cell_value(ws, row_index, col, None)
    row_index += 1

    # Total row (overwrite position in template)
    if total_row is not None:
        total_calls = int(total_row.get("Total Calls") or 0)
        total_answered = int(total_row.get("Answered") or 0)
        total_missed = int(total_row.get("Missed") or 0)
        total_abandoned = int(total_row.get("Abandoned") or 0)
        total_missed_abandoned = total_missed + total_abandoned
    else:
        total_calls = total_calls_from_branches
        total_answered = None
        total_missed_abandoned = None

    set_cell_value(ws, row_index, 1, None)
    set_cell_value(ws, row_index, 2, "Total")
    set_cell_value(ws, row_index, 3, total_calls)
    set_cell_value(ws, row_index, 4, total_answered)
    set_cell_value(ws, row_index, 5, total_missed_abandoned)
    for col in range(6, 14):
        set_cell_value(ws, row_index, col, None)

    wb.save(output_path)


def build_details_workbook(detail_csv, output_path):
    """
    Build a clean details workbook from the raw detailed CSV.
    Uses the same normalisation logic as the summary.
    """
    raw_rows = read_csv_rows(detail_csv)
    detail_rows = normalize_detail_rows(raw_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Call Center Details"

    headers = [
        "Queue",
        "Answered",
        "Missed",
        "Abandoned / Agent",
        "Status",
        "Ring Duration",
        "Answered Wait",
        "All Calls Wait",
        "Talking Time",
    ]
    ws.append(headers)

    current_queue = None

    for row in detail_rows:
        queue_cell = (row.get("Queue") or "").strip()
        total_calls = row.get("Total Calls") or ""
        answered = row.get("Answered") or ""
        missed = row.get("Missed") or ""
        abandoned = row.get("Abandoned") or ""
        status = row.get("AVG Handle Time") or ""
        ring = row.get("AVG Waiting Time (Answered Calls)") or ""
        all_wait = row.get("AVG Waiting Time (All Calls)") or ""
        talking = row.get("Average Talking Time") or ""

        # Summary header per queue
        if queue_cell:
            current_queue = queue_cell
            continue

        # Skip internal header rows
        if total_calls == "ID" or answered == "Time" or missed == "Call From":
            continue

        if not answered:
            continue

        ws.append([
            current_queue or "",
            answered,
            missed,
            abandoned,
            status,
            ring,
            ring,
            all_wait,
            talking,
        ])

    # Column widths for readability
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16

    wb.save(output_path)


def main():
    if len(sys.argv) != 6:
        print("Usage: python call_center_report.py summary.csv details.csv template_summary.xlsx out_summary.xlsx out_details.xlsx")
        sys.exit(1)

    summary_csv = sys.argv[1]
    details_csv = sys.argv[2]
    template_summary = sys.argv[3]
    out_summary = sys.argv[4]
    out_details = sys.argv[5]

    fill_summary_template(summary_csv, details_csv, template_summary, out_summary)
    build_details_workbook(details_csv, out_details)
    print("Done. Generated:")
    print(f"  {out_summary}")
    print(f"  {out_details}")


if __name__ == "__main__":
    main()
